import json
import requests
import pandas as pd
import time
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from datetime import datetime
import os
import concurrent.futures
import threading
from requests.adapters import HTTPAdapter
from concurrent.futures import ThreadPoolExecutor
import csv

class UltraFastBulkDataExtraction:
    def __init__(self, csv_file_path, output_file="bulk_drug_data.xlsx"):
        self.csv_file_path = csv_file_path
        self.output_file = output_file
        self.processed_count = 0
        self.failed_urls = []
        self.session = requests.Session()
        self.lock = threading.Lock()
        
        # Optimized session headers
        self.session.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Connection": "keep-alive"
        })
        
        # Connection adapter
        adapter = HTTPAdapter(pool_connections=30, pool_maxsize=30)
        self.session.mount('http://', adapter)
        self.session.mount('https://', adapter)
        
        print("✅ UltraFastBulkDataExtraction initialized")
        
        
    def get_current_excel_row_count(self):
        """Get current row count from existing Excel file"""
        try:
            if os.path.exists(self.output_file):
                wb = load_workbook(self.output_file, read_only=True)
                ws = wb.active
                row_count = ws.max_row - 1  # Subtract 1 for header row
                wb.close()
                print(f"📊 Found existing Excel file with {row_count} rows")
                return row_count
            return 0
        except Exception as e:
            print(f"❌ Error reading Excel file: {str(e)}")
            return 0

    def get_soup(self, url):
        """Get soup object with error handling"""
        try:
            response = self.session.get(url, timeout=8)
            response.raise_for_status()
            return BeautifulSoup(response.content, 'html.parser')
        except Exception as e:
            return None

    def extract_single_url_data(self, url):
        """Extract all data for a single URL"""
        soup = self.get_soup(url)
        if not soup:
            return None
            
        try:
            data = {
                'url': url,
                'prescription': self.prescription(soup),
                'salt_composition': self.salt_compo(soup),
                'side_effects': self.side_effects(soup),
                'product_description': self.product_description(soup),
                'faqs': self.format_faqs_for_excel(self.faqs(soup)),
                'how_drug_works': self.format_how_drug_works_for_excel(self.how_drug_works(soup)),
                'drug_interactions': self.format_drug_interactions_for_excel(self.drug_interaction(soup)),
                'how_to_use': self.format_how_to_use_for_excel(self.how_to_use(soup)),
                'safety_advice': self.format_safety_advice_for_excel(self.safety_advice(soup))
            }
            return data
        except Exception as e:
            print(f"❌ Error extracting data from {url}: {str(e)}")
            return None

    def prescription(self, soup):
        try:
            outer_div = soup.find("div", class_="DrugHeader__prescription-req___34WVy")
            if outer_div:
                span = outer_div.find("span")
                if span:
                    return span.get_text(strip=True)
        except:
            pass
        return ""

    def salt_compo(self, soup):
        try:
            div_tag = soup.find("div", class_="saltInfo DrugHeader__meta-value___vqYM0")
            if div_tag:
                a_tag = div_tag.find("a")
                if a_tag:
                    return a_tag.get_text(strip=True)
        except:
            pass
        return ""

    def side_effects(self, soup):
        try:
            side_effects_div = soup.find("div", id="side_effects")
            if side_effects_div:
                return side_effects_div.get_text(separator="\n", strip=True)
        except:
            pass
        return ""

    def product_description(self, soup):
        try:
            product_intro_div = soup.find("div", class_="DrugOverview__container___CqA8x")
            if product_intro_div:
                return product_intro_div.get_text(separator=" ", strip=True)
        except:
            pass
        return ""

    def faqs(self, soup):
        faq_list = []
        try:
            faq_div = soup.find("div", id="faq")
            if faq_div:
                for tile in faq_div.find_all("div", class_="Faqs__tile___1B58W"):
                    question_tag = tile.find("h3", class_="Faqs__ques___1iPB9")
                    answer_tag = tile.find("div", class_="Faqs__ans___1uuIW")
                    if question_tag and answer_tag:
                        faq_list.append({
                            "Q": question_tag.get_text(strip=True),
                            "A": answer_tag.get_text(strip=True)
                        })
        except:
            pass
        return faq_list

    def how_drug_works(self, soup):
        try:
            container = soup.find("div", id="how_drug_works")
            if container:
                title_tag = container.find("h2")
                content_tag = container.find("div", class_="DrugOverview__content___22ZBX")
                title = title_tag.get_text(strip=True) if title_tag else ""
                content = content_tag.get_text(strip=True) if content_tag else ""
                return {"title": title, "content": content}
        except:
            pass
        return {}

    def drug_interaction(self, soup):
        try:
            interaction_div = soup.find("div", id="drug_interaction")
            if interaction_div:
                title_tag = interaction_div.find("h2")
                description_tag = interaction_div.find("div", class_="DrugInteraction__desc___2y8bR")
                title = title_tag.get_text(strip=True) if title_tag else ""
                description = description_tag.get_text(strip=True) if description_tag else ""

                interactions = []
                for drug_block in interaction_div.find_all("div", class_="DrugInteraction__drug___1XyzI"):
                    name = drug_block.get_text(strip=True)
                    desc_block = drug_block.find_next("div", class_="DrugInteraction__interaction-text___1hOwx")
                    desc_text = desc_block.get_text(" ", strip=True) if desc_block else ""
                    interactions.append({"drug": name, "description": desc_text})

                return {
                    "title": title,
                    "description": description,
                    "interactions": interactions
                }
        except:
            pass
        return {}

    def how_to_use(self, soup):
        try:
            how_to_use_div = soup.find("div", id="how_to_use")
            if how_to_use_div:
                heading_tag = how_to_use_div.find("h2")
                instructions_tag = how_to_use_div.find("div", class_="DrugOverview__content___22ZBX")
                heading = heading_tag.get_text(strip=True) if heading_tag else ""
                instructions = instructions_tag.get_text(" ", strip=True) if instructions_tag else ""
                return {"heading": heading, "instructions": instructions}
        except:
            pass
        return {}

    def safety_advice(self, soup):
        try:
            safety_div = soup.find("div", id="safety_advice")
            if not safety_div:
                return {}

            heading_tag = safety_div.find("h2")
            heading = heading_tag.get_text(strip=True) if heading_tag else ""

            safety_items = []
            warnings = safety_div.find_all("div", class_="DrugOverview__warning-top___UD3xX")

            for warning in warnings:
                label_tag = warning.find("span")
                status_tag = warning.find("div", class_="DrugOverview__warning-tag___aHZlc")

                label = label_tag.get_text(strip=True) if label_tag else ""
                status = status_tag.get_text(strip=True) if status_tag else ""

                desc_div = warning.find_next_sibling("div", class_="DrugOverview__content___22ZBX")
                description = desc_div.get_text(" ", strip=True) if desc_div else ""

                safety_items.append({
                    "label": label,
                    "status": status,
                    "description": description
                })

            return {"heading": heading, "items": safety_items}
        except:
            pass
        return {}

    def format_faqs_for_excel(self, faqs):
        if not faqs:
            return ""
        formatted_text = ""
        for i, faq in enumerate(faqs, 1):
            formatted_text += f"Q{i}: {faq['Q']}\n"
            formatted_text += f"A{i}: {faq['A']}\n\n"
        return formatted_text.strip()

    def format_drug_interactions_for_excel(self, interactions_data):
        if not interactions_data:
            return ""
        formatted_text = ""
        if interactions_data.get('title'):
            formatted_text += f"TITLE: {interactions_data['title']}\n\n"
        if interactions_data.get('description'):
            formatted_text += f"DESCRIPTION: {interactions_data['description']}\n\n"
        if interactions_data.get('interactions'):
            formatted_text += "DRUG INTERACTIONS:\n"
            for i, interaction in enumerate(interactions_data['interactions'], 1):
                formatted_text += f"{i}. {interaction['drug']}\n"
                formatted_text += f"   Description: {interaction['description']}\n\n"
        return formatted_text.strip()

    def format_safety_advice_for_excel(self, safety_data):
        if not safety_data:
            return ""
        formatted_text = ""
        if safety_data.get('heading'):
            formatted_text += f"HEADING: {safety_data['heading']}\n\n"
        if safety_data.get('items'):
            for i, item in enumerate(safety_data['items'], 1):
                formatted_text += f"{i}. {item['label']}\n"
                formatted_text += f"   Status: {item['status']}\n"
                formatted_text += f"   Description: {item['description']}\n\n"
        return formatted_text.strip()

    def format_how_drug_works_for_excel(self, how_works_data):
        if not how_works_data:
            return ""
        formatted_text = ""
        if how_works_data.get('title'):
            formatted_text += f"TITLE: {how_works_data['title']}\n\n"
        if how_works_data.get('content'):
            formatted_text += f"CONTENT: {how_works_data['content']}"
        return formatted_text.strip()

    def format_how_to_use_for_excel(self, how_to_use_data):
        if not how_to_use_data:
            return ""
        formatted_text = ""
        if how_to_use_data.get('heading'):
            formatted_text += f"HEADING: {how_to_use_data['heading']}\n\n"
        if how_to_use_data.get('instructions'):
            formatted_text += f"INSTRUCTIONS: {how_to_use_data['instructions']}"
        return formatted_text.strip()

    def create_excel_with_headers(self):
        """Create Excel file with headers if it doesn't exist"""
        if os.path.exists(self.output_file):
            print(f"📊 Excel file already exists: {self.output_file}")
            return
            
        wb = Workbook()
        ws = wb.active
        ws.title = "Drug Data"

        headers = [
            "URL",
            "Prescription Required",
            "Salt Composition", 
            "Side Effects",
            "Product Description",
            "FAQs",
            "How Drug Works",
            "Drug Interactions",
            "How to Use",
            "Safety Advice"
        ]
        
        ws.append(headers)
        
        # Format headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        column_widths = [50, 20, 30, 40, 60, 60, 60, 60, 60, 60]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        wb.save(self.output_file)
        print(f"✅ Created Excel file: {self.output_file}")

    def append_data_to_excel_batch(self, data_batch):
        """Append batch of data to existing Excel file"""
        with self.lock:
            try:
                wb = load_workbook(self.output_file)
                ws = wb.active
                
                for data in data_batch:
                    if data:
                        row_data = [
                            data['url'],
                            data['prescription'],
                            data['salt_composition'],
                            data['side_effects'],
                            data['product_description'],
                            data['faqs'],
                            data['how_drug_works'],
                            data['drug_interactions'],
                            data['how_to_use'],
                            data['safety_advice']
                        ]
                        ws.append(row_data)
                        
                        row_num = ws.max_row
                        for cell in ws[row_num]:
                            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                        ws.row_dimensions[row_num].height = 100
                
                wb.save(self.output_file)
                current_rows = ws.max_row - 1
                print(f"✅ Saved batch to Excel. Total rows: {current_rows}")
                
            except Exception as e:
                print(f"❌ Error saving to Excel: {str(e)}")

    def process_urls_chunk(self, urls_chunk):
        """Process chunk of URLs"""
        results = []
        for url in urls_chunk:
            try:
                data = self.extract_single_url_data(url.strip())
                if data:
                    results.append(data)
                    with self.lock:
                        self.processed_count += 1
                        if self.processed_count % 10 == 0:  # Progress every 10
                            print(f"⚡ Processed: {self.processed_count} URLs")
                else:
                    with self.lock:
                        self.failed_urls.append(url)
            except Exception as e:
                with self.lock:
                    self.failed_urls.append(url)
                    print(f"⚠️ Error processing {url}: {str(e)}")
        return results

    def create_csv_with_headers(self, csv_file):
        """Create CSV file with headers if it doesn't exist"""
        if os.path.exists(csv_file):
            print(f"📊 CSV file already exists: {csv_file}")
            return
        headers = [
            "URL",
            "Prescription Required",
            "Salt Composition", 
            "Side Effects",
            "Product Description",
            "FAQs",
            "How Drug Works",
            "Drug Interactions",
            "How to Use",
            "Safety Advice"
        ]
        with open(csv_file, mode='w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
        print(f"✅ Created CSV file: {csv_file}")

    def append_data_to_csv_batch(self, data_batch, csv_file):
        """Append batch of data to CSV file"""
        with self.lock:
            try:
                with open(csv_file, mode='a', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    for data in data_batch:
                        if data:
                            row_data = [
                                data['url'],
                                data['prescription'],
                                data['salt_composition'],
                                data['side_effects'],
                                data['product_description'],
                                data['faqs'],
                                data['how_drug_works'],
                                data['drug_interactions'],
                                data['how_to_use'],
                                data['safety_advice']
                            ]
                            writer.writerow(row_data)
                print(f"✅ Saved batch to CSV. Total rows written: {self.processed_count}")
            except Exception as e:
                print(f"❌ Error saving to CSV: {str(e)}")

    def ultra_fast_process(self, start_from=114990, batch_size=100, max_workers=10, limit=30000, csv_file="bulk_drug_114990_30k.csv"):
        """Ultra-fast processing with threading, saving to CSV"""
        print(f"🚀 Starting ultra-fast processing from row {start_from + 1}")

        # Read CSV efficiently
        try:
            print(f"📊 Reading CSV file: {self.csv_file_path}")
            df = pd.read_csv(self.csv_file_path, header=None)
            all_urls = df[0].dropna().tolist()
            # Skip to start_from position and limit to next 30,000
            urls = all_urls[start_from:start_from+limit]
            total_urls = len(urls)
            print(f"📈 Loaded {total_urls} URLs to process (starting from row {start_from + 1})")
        except Exception as e:
            print(f"❌ Error reading CSV: {str(e)}")
            return

        # Create CSV file if needed
        if not os.path.exists(csv_file):
            self.create_csv_with_headers(csv_file)

        print(f"⚡ Processing with {max_workers} threads, batch size: {batch_size}")

        start_time = time.time()
        save_counter = 0
        pending_data = []

        # Process in batches
        for i in range(0, total_urls, batch_size):
            batch_urls = urls[i:i+batch_size]
            batch_num = i // batch_size + 1
            total_batches = (total_urls + batch_size - 1) // batch_size

            print(f"\n🔥 Processing Batch {batch_num}/{total_batches} ({len(batch_urls)} URLs)")
            batch_start = time.time()

            # Split into chunks for threading
            chunk_size = max(1, len(batch_urls) // max_workers)
            chunks = [batch_urls[j:j+chunk_size] for j in range(0, len(batch_urls), chunk_size)]

            batch_results = []
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                future_to_chunk = {executor.submit(self.process_urls_chunk, chunk): chunk for chunk in chunks}
                for future in concurrent.futures.as_completed(future_to_chunk):
                    try:
                        chunk_results = future.result(timeout=120)
                        batch_results.extend(chunk_results)
                    except Exception as e:
                        print(f"⚠️ Chunk error: {str(e)}")

            # Add to pending data
            pending_data.extend(batch_results)
            save_counter += len(batch_results)

            # Save every 100 successful records
            if save_counter >= 100 or i + batch_size >= total_urls:
                if pending_data:
                    self.append_data_to_csv_batch(pending_data, csv_file)
                    print(f"💾 Saved {len(pending_data)} records to CSV")
                    pending_data = []
                    save_counter = 0

            # Progress stats
            batch_time = time.time() - batch_start
            print(f"⏱️ Batch completed in {batch_time:.1f}s")
            if batch_time > 0:
                print(f"🎯 Batch speed: {len(batch_urls)/batch_time:.1f} URLs/sec")

            # ETA calculation
            if i + batch_size < total_urls:
                elapsed = time.time() - start_time
                processed = i + batch_size
                remaining = total_urls - processed
                if processed > 0:
                    avg_time_per_url = elapsed / processed
                    eta_seconds = remaining * avg_time_per_url
                    print(f"🕐 ETA: {eta_seconds/60:.1f} minutes")

        # Final statistics
        total_time = time.time() - start_time
        print("\n" + "="*60)
        print("🎉 PROCESSING COMPLETE!")
        print(f"📊 Total URLs processed: {self.processed_count}")
        print(f"❌ Failed URLs: {len(self.failed_urls)}")
        if self.processed_count + len(self.failed_urls) > 0:
            success_rate = (self.processed_count/(self.processed_count + len(self.failed_urls)))*100
            print(f"✅ Success rate: {success_rate:.1f}%")
        print(f"⏱️ Total time: {total_time/60:.1f} minutes")
        if total_time > 0:
            print(f"⚡ Average speed: {self.processed_count/total_time:.2f} URLs/second")
        print(f"💾 CSV file: {csv_file}")

        # Save failed URLs
        if self.failed_urls:
            failed_df = pd.DataFrame(self.failed_urls, columns=['failed_urls'])
            failed_file = f"failed_urls_from_{start_from}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            failed_df.to_csv(failed_file, index=False)
            print(f"📝 Failed URLs saved to: {failed_file}")

    def csv_to_excel(self, csv_file, excel_file):
        """Convert CSV file to Excel file"""
        df = pd.read_csv(csv_file)
        df.to_excel(excel_file, index=False)
        print(f"✅ Converted {csv_file} to {excel_file}")

# Main execution
if __name__ == "__main__":
    print("🚀 Starting UltraFast Bulk Data Extraction...")
    print("📍 Starting from row 174990 in CSV, processing all remaining URLs, saving to CSV.")

    csv_file = "URL.csv"
    output_csv = "bulk_drug_174990_to_end.csv"
    output_excel = "bulk_drug_174990_to_end.xlsx"

    # Check if CSV file exists
    if not os.path.exists(csv_file):
        print(f"❌ CSV file not found: {csv_file}")
        print("Please make sure the CSV file is in the same directory as this script.")
        exit(1)

    # Count total URLs to determine limit
    df = pd.read_csv(csv_file, header=None)
    total_urls = len(df)
    start_from = 174990
    limit = total_urls - start_from

    extractor = UltraFastBulkDataExtraction(
        csv_file_path=csv_file,
        output_file=output_csv  # Not used for CSV, but kept for compatibility
    )
    extractor.create_csv_with_headers(output_csv)  # Always create new CSV file with headers

    # Start processing from row 174990, process all remaining URLs, save to CSV
    extractor.ultra_fast_process(
        start_from=start_from,
        batch_size=100,
        max_workers=10,
        limit=limit,
        csv_file=output_csv
    )

    # After processing, convert CSV to Excel
    extractor.csv_to_excel(output_csv, output_excel)