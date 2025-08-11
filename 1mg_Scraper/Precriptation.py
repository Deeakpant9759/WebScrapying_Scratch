import json
import requests
import pandas as pd
import time
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from datetime import datetime
import os
import concurrent.futures
import threading
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import random

class SuperFastBulkDataExtraction:
    def __init__(self, csv_file_path, output_file="bulk_drug_data.xlsx"):
        self.csv_file_path = csv_file_path
        self.output_file = output_file
        self.processed_count = 0
        self.failed_urls = []
        self.lock = threading.Lock()
        
        # Large pool of real user agents (more reliable than fake-useragent library)
        self.user_agents = [
            # Chrome Windows
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Safari/537.36",
            # Chrome Mac
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            # Chrome Linux
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
            "Mozilla/5.0 (X11; Ubuntu; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            # Firefox Windows
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:120.0) Gecko/20100101 Firefox/120.0",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:119.0) Gecko/20100101 Firefox/119.0",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:118.0) Gecko/20100101 Firefox/118.0",
            # Firefox Mac
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:120.0) Gecko/20100101 Firefox/120.0",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:119.0) Gecko/20100101 Firefox/119.0",
            # Firefox Linux
            "Mozilla/5.0 (X11; Linux x86_64; rv:120.0) Gecko/20100101 Firefox/120.0",
            "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:120.0) Gecko/20100101 Firefox/120.0",
            # Edge
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Edge/120.0.0.0",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Edge/119.0.0.0",
            # Safari
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.1 Safari/605.1.15",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.6 Safari/605.1.15",
            # Mobile user agents for variety
            "Mozilla/5.0 (iPhone; CPU iPhone OS 17_1 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.1 Mobile/15E148 Safari/604.1",
            "Mozilla/5.0 (Linux; Android 13; SM-G991B) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Mobile Safari/537.36",
            "Mozilla/5.0 (Linux; Android 13; Pixel 7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Mobile Safari/537.36",
            # Additional Chrome versions
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36",
            # Additional Firefox versions
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:117.0) Gecko/20100101 Firefox/117.0",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:116.0) Gecko/20100101 Firefox/116.0",
            # Opera
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36 OPR/106.0.0.0",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36 OPR/106.0.0.0"
        ]
        
        self.session = requests.Session()
        print(f"✅ Loaded {len(self.user_agents)} user agents for rotation")
        
        # Aggressive retry strategy
        retry_strategy = Retry(
            total=2,  # Reduced retries for speed
            status_forcelist=[429, 500, 502, 503, 504],
            backoff_factor=0.1  # Faster backoff
        )
        
        # Enhanced connection pooling
        adapter = HTTPAdapter(
            pool_connections=50,  # Increased
            pool_maxsize=50,      # Increased
            max_retries=retry_strategy,
            pool_block=False
        )
        
        self.session.mount('http://', adapter)
        self.session.mount('https://', adapter)
        
        # Keep-alive and connection reuse
        self.session.keep_alive = True
        
    def get_random_user_agent(self):
        """Get truly random user agent from pool"""
        return random.choice(self.user_agents)
    
    def update_session_headers(self):
        """Update session with rotated user agent"""
        self.session.headers.update({
            "User-Agent": self.get_random_user_agent(),
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.5",
            "Accept-Encoding": "gzip, deflate, br",
            "Accept-Charset": "utf-8, iso-8859-1;q=0.5",  # Handle encoding
            "DNT": "1",
            "Connection": "keep-alive",
            "Upgrade-Insecure-Requests": "1",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Cache-Control": "max-age=0"
        })
        
    def get_current_excel_row_count(self):
        """Get current row count from existing Excel file - optimized"""
        try:
            if os.path.exists(self.output_file):
                wb = load_workbook(self.output_file, read_only=True, data_only=True)
                ws = wb.active
                row_count = ws.max_row - 1  # Subtract 1 for header row
                wb.close()
                return row_count
            return 0
        except Exception as e:
            print(f"❌ Error reading Excel file: {str(e)}")
            return 0

    def get_soup(self, url):
        """Encoding-safe soup retrieval with fake user agents"""
        try:
            # Get fresh random user agent for each request
            headers = {
                "User-Agent": self.get_random_user_agent(),
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
                "Accept-Language": "en-US,en;q=0.5",
                "Accept-Encoding": "gzip, deflate, br",
                "Accept-Charset": "utf-8, iso-8859-1;q=0.5",
                "DNT": "1",
                "Connection": "keep-alive",
                "Cache-Control": "max-age=0",
                "Sec-Fetch-Dest": "document",
                "Sec-Fetch-Mode": "navigate",
                "Sec-Fetch-Site": "none"
            }
            
            # Get response with proper encoding handling
            response = self.session.get(url, timeout=5, stream=False, headers=headers)
            response.raise_for_status()
            
            # Fix encoding issues
            if response.encoding is None or response.encoding.lower() in ['iso-8859-1', 'windows-1252']:
                response.encoding = 'utf-8'
            
            # Get content with encoding fallbacks
            try:
                content = response.content
                # Try to decode and re-encode to fix encoding issues
                if response.text:
                    content = response.text.encode('utf-8', errors='ignore')
            except:
                content = response.content
            
            # Parse with encoding error handling
            try:
                # Use lxml with encoding handling
                return BeautifulSoup(content, 'lxml', from_encoding='utf-8')
            except:
                try:
                    # Fallback to html.parser with error handling
                    return BeautifulSoup(content, 'html.parser', from_encoding='utf-8')
                except:
                    # Last resort - ignore encoding errors completely
                    clean_content = content.decode('utf-8', errors='ignore').encode('utf-8')
                    return BeautifulSoup(clean_content, 'html.parser')
                
        except requests.exceptions.Timeout:
            return None  # Silent timeout for speed
        except requests.exceptions.RequestException:
            return None  # Silent request errors for speed
        except Exception:
            return None  # Silent other errors for speed

    def extract_single_url_data(self, url):
        """Optimized data extraction with error handling"""
        soup = self.get_soup(url)
        if not soup:
            return None
            
        try:
            # Pre-find common elements to avoid repeated searches
            data = {
                'url': url,
                'prescription': self.prescription(soup),
                'salt_composition': self.salt_compo(soup),
                'side_effects': self.side_effects(soup),
                'product_description': self.product_description(soup),
                'faqs': self.format_faqs_for_excel(self.faqs(soup)),
                'how_drug_works': self.format_how_drug_works_for_excel(self.how_drug_works(soup)),
                'drug_interactions': self.format_drug_interactions_for_excel(self.drug_interaction(soup)),
                'how_to_use': self.format_how_to_use_for_excel(self.how_to_use(soup)),
                'safety_advice': self.format_safety_advice_for_excel(self.safety_advice(soup))
            }
            
            with self.lock:
                self.processed_count += 1
                # Reduced progress reporting frequency for speed
                if self.processed_count % 50 == 0:  
                    print(f"🔄 Processed: {self.processed_count}")
                    
            return data
            
        except Exception as e:
            print(f"❌ Data extraction error {url}: {str(e)[:50]}")
            return None

    # Optimized extraction methods with minimal DOM traversal
    def prescription(self, soup):
        try:
            outer_div = soup.find("div", class_="DrugHeader__prescription-req___34WVy")
            return outer_div.find("span").get_text(strip=True) if outer_div and outer_div.find("span") else ""
        except:
            return ""

    def salt_compo(self, soup):
        try:
            div_tag = soup.find("div", class_="saltInfo DrugHeader__meta-value___vqYM0")
            a_tag = div_tag.find("a") if div_tag else None
            return a_tag.get_text(strip=True) if a_tag else ""
        except:
            return ""

    def side_effects(self, soup):
        try:
            side_effects_div = soup.find("div", id="side_effects")
            return side_effects_div.get_text(separator="\n", strip=True) if side_effects_div else ""
        except:
            return ""

    def product_description(self, soup):
        try:
            product_intro_div = soup.find("div", class_="DrugOverview__container___CqA8x")
            return product_intro_div.get_text(separator=" ", strip=True) if product_intro_div else ""
        except:
            return ""

    def faqs(self, soup):
        try:
            faq_div = soup.find("div", id="faq")
            if not faq_div:
                return []
                
            faq_list = []
            tiles = faq_div.find_all("div", class_="Faqs__tile___1B58W")
            
            for tile in tiles:
                question_tag = tile.find("h3", class_="Faqs__ques___1iPB9")
                answer_tag = tile.find("div", class_="Faqs__ans___1uuIW")
                if question_tag and answer_tag:
                    faq_list.append({
                        "Q": question_tag.get_text(strip=True),
                        "A": answer_tag.get_text(strip=True)
                    })
            return faq_list
        except:
            return []

    def how_drug_works(self, soup):
        try:
            container = soup.find("div", id="how_drug_works")
            if not container:
                return {}
                
            title_tag = container.find("h2")
            content_tag = container.find("div", class_="DrugOverview__content___22ZBX")
            
            return {
                "title": title_tag.get_text(strip=True) if title_tag else "",
                "content": content_tag.get_text(strip=True) if content_tag else ""
            }
        except:
            return {}

    def drug_interaction(self, soup):
        try:
            interaction_div = soup.find("div", id="drug_interaction")
            if not interaction_div:
                return {}

            title_tag = interaction_div.find("h2")
            description_tag = interaction_div.find("div", class_="DrugInteraction__desc___2y8bR")
            
            interactions = []
            drug_blocks = interaction_div.find_all("div", class_="DrugInteraction__drug___1XyzI")
            
            for drug_block in drug_blocks:
                name = drug_block.get_text(strip=True)
                desc_block = drug_block.find_next("div", class_="DrugInteraction__interaction-text___1hOwx")
                desc_text = desc_block.get_text(" ", strip=True) if desc_block else ""
                interactions.append({"drug": name, "description": desc_text})

            return {
                "title": title_tag.get_text(strip=True) if title_tag else "",
                "description": description_tag.get_text(strip=True) if description_tag else "",
                "interactions": interactions
            }
        except:
            return {}

    def how_to_use(self, soup):
        try:
            how_to_use_div = soup.find("div", id="how_to_use")
            if not how_to_use_div:
                return {}
                
            heading_tag = how_to_use_div.find("h2")
            instructions_tag = how_to_use_div.find("div", class_="DrugOverview__content___22ZBX")
            
            return {
                "heading": heading_tag.get_text(strip=True) if heading_tag else "",
                "instructions": instructions_tag.get_text(" ", strip=True) if instructions_tag else ""
            }
        except:
            return {}

    def safety_advice(self, soup):
        try:
            safety_div = soup.find("div", id="safety_advice")
            if not safety_div:
                return {}

            heading_tag = safety_div.find("h2")
            heading = heading_tag.get_text(strip=True) if heading_tag else ""

            safety_items = []
            warnings = safety_div.find_all("div", class_="DrugOverview__warning-top___UD3xX")

            for warning in warnings:
                label_tag = warning.find("span")
                status_tag = warning.find("div", class_="DrugOverview__warning-tag___aHZlc")
                desc_div = warning.find_next_sibling("div", class_="DrugOverview__content___22ZBX")

                safety_items.append({
                    "label": label_tag.get_text(strip=True) if label_tag else "",
                    "status": status_tag.get_text(strip=True) if status_tag else "",
                    "description": desc_div.get_text(" ", strip=True) if desc_div else ""
                })

            return {"heading": heading, "items": safety_items}
        except:
            return {}

    # Format methods remain the same but with try-catch
    def format_faqs_for_excel(self, faqs):
        try:
            if not faqs:
                return ""
            formatted_text = ""
            for i, faq in enumerate(faqs, 1):
                formatted_text += f"Q{i}: {faq['Q']}\n"
                formatted_text += f"A{i}: {faq['A']}\n\n"
            return formatted_text.strip()
        except:
            return ""

    def format_drug_interactions_for_excel(self, interactions_data):
        try:
            if not interactions_data:
                return ""
            formatted_text = ""
            if interactions_data.get('title'):
                formatted_text += f"TITLE: {interactions_data['title']}\n\n"
            if interactions_data.get('description'):
                formatted_text += f"DESCRIPTION: {interactions_data['description']}\n\n"
            if interactions_data.get('interactions'):
                formatted_text += "DRUG INTERACTIONS:\n"
                for i, interaction in enumerate(interactions_data['interactions'], 1):
                    formatted_text += f"{i}. {interaction['drug']}\n"
                    formatted_text += f"   Description: {interaction['description']}\n\n"
            return formatted_text.strip()
        except:
            return ""

    def format_safety_advice_for_excel(self, safety_data):
        try:
            if not safety_data:
                return ""
            formatted_text = ""
            if safety_data.get('heading'):
                formatted_text += f"HEADING: {safety_data['heading']}\n\n"
            if safety_data.get('items'):
                for i, item in enumerate(safety_data['items'], 1):
                    formatted_text += f"{i}. {item['label']}\n"
                    formatted_text += f"   Status: {item['status']}\n"
                    formatted_text += f"   Description: {item['description']}\n\n"
            return formatted_text.strip()
        except:
            return ""

    def format_how_drug_works_for_excel(self, how_works_data):
        try:
            if not how_works_data:
                return ""
            formatted_text = ""
            if how_works_data.get('title'):
                formatted_text += f"TITLE: {how_works_data['title']}\n\n"
            if how_works_data.get('content'):
                formatted_text += f"CONTENT: {how_works_data['content']}"
            return formatted_text.strip()
        except:
            return ""

    def format_how_to_use_for_excel(self, how_to_use_data):
        try:
            if not how_to_use_data:
                return ""
            formatted_text = ""
            if how_to_use_data.get('heading'):
                formatted_text += f"HEADING: {how_to_use_data['heading']}\n\n"
            if how_to_use_data.get('instructions'):
                formatted_text += f"INSTRUCTIONS: {how_to_use_data['instructions']}"
            return formatted_text.strip()
        except:
            return ""

    def create_excel_with_headers(self):
        """Create Excel file with headers if it doesn't exist"""
        if os.path.exists(self.output_file):
            return
            
        wb = Workbook()
        ws = wb.active
        ws.title = "Drug Data"

        headers = [
            "URL", "Prescription Required", "Salt Composition", "Side Effects",
            "Product Description", "FAQs", "How Drug Works", "Drug Interactions",
            "How to Use", "Safety Advice"
        ]
        
        ws.append(headers)
        
        # Format headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        column_widths = [50, 20, 30, 40, 60, 60, 60, 60, 60, 60]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        wb.save(self.output_file)
        print(f"✅ Created Excel file: {self.output_file}")

    def append_data_to_excel(self, data_batch):
        """Optimized batch append to Excel"""
        with self.lock:
            try:
                wb = load_workbook(self.output_file)
                ws = wb.active
                
                for data in data_batch:
                    if data:
                        row_data = [
                            data['url'], data['prescription'], data['salt_composition'],
                            data['side_effects'], data['product_description'], data['faqs'],
                            data['how_drug_works'], data['drug_interactions'], 
                            data['how_to_use'], data['safety_advice']
                        ]
                        ws.append(row_data)
                        
                        # Minimal formatting for speed
                        row_num = ws.max_row
                        for cell in ws[row_num]:
                            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                        ws.row_dimensions[row_num].height = 100
                
                wb.save(self.output_file)
                total_rows = ws.max_row - 1
                print(f"💾 Saved batch. Total: {total_rows} rows")
                
            except Exception as e:
                print(f"❌ Excel save error: {str(e)}")

    def process_urls_batch(self, urls_batch):
        """Process batch with better error handling"""
        results = []
        for url in urls_batch:
            try:
                data = self.extract_single_url_data(url)
                if data:
                    results.append(data)
                else:
                    with self.lock:
                        self.failed_urls.append(url)
            except Exception as e:
                with self.lock:
                    self.failed_urls.append(url)
                    print(f"❌ Batch processing error: {str(e)[:50]}")
        return results

    def process_urls_from_csv_super_fast(self, start_batch=131, has_header=False, batch_size=200, max_workers=12):
        """SUPER FAST processing with optimizations - Continue from specific batch"""
        
        # Calculate starting row from batch number
        start_from_row = (start_batch - 1) * batch_size
        
        print(f"🚀 SUPER FAST MODE: Continuing from BATCH {start_batch}")
        print(f"📍 Starting from ROW {start_from_row}")
        
        # Read CSV efficiently
        try:
            if has_header:
                df = pd.read_csv(self.csv_file_path)
                url_column = next((col for col in ['url', 'URL', 'urls', 'URLs', 'link', 'links'] if col in df.columns), df.columns[0])
                urls = df[url_column].tolist()
                print(f"📊 Using column '{url_column}': {len(df)} URLs loaded")
            else:
                df = pd.read_csv(self.csv_file_path, header=None)
                urls = df[0].tolist()
                print(f"📊 {len(df)} URLs loaded from CSV")
                
        except Exception as e:
            print(f"❌ CSV read error: {str(e)}")
            return

        # Verify Excel file exists with expected rows
        current_excel_rows = self.get_current_excel_row_count()
        print(f"📋 Current Excel file has: {current_excel_rows} rows")
        
        if current_excel_rows != start_from_row:
            print(f"⚠️  WARNING: Excel has {current_excel_rows} rows but expected {start_from_row} rows")
            print(f"⚠️  Adjusting start position to match Excel file...")
            start_from_row = current_excel_rows
            start_batch = (start_from_row // batch_size) + 1

        # Create Excel if needed
        if not os.path.exists(self.output_file):
            self.create_excel_with_headers()
        
        # Process from calculated starting row
        urls = urls[start_from_row:]
        total_urls = len(urls)
        total_batches_remaining = (total_urls + batch_size - 1) // batch_size
        total_batches_overall = (len(df) + batch_size - 1) // batch_size
        
        print(f"🔥 RESUMING PROCESSING:")
        print(f"📊 Total URLs in CSV: {len(df)}")
        print(f"✅ Already processed: {start_from_row} URLs ({start_batch-1} batches)")
        print(f"🔄 Remaining: {total_urls} URLs ({total_batches_remaining} batches)")
        print(f"📈 Overall progress: {start_batch-1}/{total_batches_overall} batches ({((start_batch-1)/total_batches_overall)*100:.1f}%)")
        print(f"⚡ Settings: Batch size={batch_size}, Workers={max_workers}")
        print(f"🕐 Estimated time: {(total_urls/batch_size/max_workers)*3:.1f} minutes")
        print("="*60)
        
        start_time = time.time()
        
        # Process in larger batches with more workers
        for i in range(0, total_urls, batch_size):
            batch_urls = urls[i:i+batch_size]
            current_batch_num = i // batch_size + 1
            overall_batch_num = start_batch + (i // batch_size)  # Overall batch number
            
            batch_start = time.time()
            
            # Parallel processing
            batch_results = []
            with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
                chunk_size = max(1, len(batch_urls) // max_workers)
                chunks = [batch_urls[j:j+chunk_size] for j in range(0, len(batch_urls), chunk_size)]
                
                future_to_chunk = {executor.submit(self.process_urls_batch, chunk): chunk for chunk in chunks}
                
                for future in concurrent.futures.as_completed(future_to_chunk):
                    try:
                        chunk_results = future.result()
                        batch_results.extend(chunk_results)
                    except Exception as e:
                        print(f"❌ Chunk error: {str(e)[:50]}")
            
            # Save results
            if batch_results:
                self.append_data_to_excel(batch_results)
            
            # Performance stats
            batch_time = time.time() - batch_start
            processed_so_far = start_from_row + min(i + batch_size, total_urls)
            total_elapsed = time.time() - start_time
            
            urls_per_second = len(batch_urls) / batch_time if batch_time > 0 else 0
            overall_progress = (overall_batch_num / total_batches_overall) * 100
            
            print(f"🏃‍♂️ BATCH {overall_batch_num}/{total_batches_overall} ({current_batch_num}/{total_batches_remaining} remaining)")
            print(f"📊 Overall progress: {overall_progress:.1f}% | Speed: {urls_per_second:.1f} URLs/sec")
            print(f"💾 Total rows in Excel: {processed_so_far}")
            
            if i + batch_size < total_urls:
                remaining_batches = total_batches_remaining - current_batch_num
                est_remaining_time = remaining_batches * batch_time
                print(f"🕐 ETA: {est_remaining_time/60:.1f} minutes remaining")
            
            print("-" * 50)
        
        # Final summary
        total_time = time.time() - start_time
        total_processed = self.processed_count
        
        print("\n" + "🎉" * 20)
        print("SUPER FAST PROCESSING COMPLETE!")
        print("🎉" * 20)
        print(f"📊 URLs processed: {total_processed}")
        print(f"❌ Failed: {len(self.failed_urls)}")
        print(f"✅ Success rate: {(total_processed/(total_processed + len(self.failed_urls)))*100:.1f}%")
        print(f"⏱️ Total time: {total_time/60:.1f} minutes")
        print(f"🚀 Average speed: {(total_processed + len(self.failed_urls))/total_time:.2f} URLs/second")
        print(f"💾 Output: {self.output_file}")
        
        # Save failed URLs
        if self.failed_urls:
            failed_file = f"failed_urls_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            failed_df = pd.DataFrame(self.failed_urls, columns=['failed_urls'])
            failed_df.to_csv(failed_file, index=False)
            print(f"📝 Failed URLs: {failed_file}")

# USAGE - OPTIMIZED FOR SPEED
if __name__ == "__main__":
    print("🚀 SUPER FAST DRUG DATA EXTRACTOR")
    print("🔥 Optimized for maximum speed!")
    
    # Your settings
    csv_file = "URL.csv"  # Your CSV file with URLs
    
    # Initialize with super fast settings
    extractor = SuperFastBulkDataExtraction(
        csv_file_path=csv_file,
        output_file="bulk_drug_data.xlsx"
    )
    
    # SUPER FAST MODE - Continue from batch 143 (current Excel rows: 28,563)
    # The script auto-detected you're at batch 142, so continuing from 143
    extractor.process_urls_from_csv_super_fast(
        start_batch=143,     # Auto-detected from your current progress
        has_header=False,    # Set to True if your CSV has headers  
        batch_size=200,      # Keep same batch size
        max_workers=16       # Increased workers to handle encoding delays
    )
    
    print("🎯 TIP: For maximum performance:")
    print("   1. Install lxml for faster parsing: pip install lxml")  
    print("   2. Use SSD storage for faster Excel I/O")
    print("   3. Close other applications to free up resources")
    print("   4. Increase max_workers if you have more CPU cores")
    print("   5. Check your internet speed - it might be the bottleneck")