import json
import requests
import pandas as pd
import time
import random
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from datetime import datetime
from fake_useragent import UserAgent
import os
import concurrent.futures
import threading
from requests.adapters import HTTPAdapter
from concurrent.futures import ThreadPoolExecutor
import csv

class EnhancedUltraFastBulkDataExtraction:
    def __init__(self, csv_file_path, output_file="bulk_drug_data.xlsx", column_index=5, save_html=True):
        self.csv_file_path = csv_file_path
        self.output_file = output_file
        self.column_index = column_index  # Column index for URLs (0-based)
        self.save_html = save_html
        self.processed_count = 0
        self.failed_urls = []
        self.lock = threading.Lock()
        
        # Initialize User Agent
        self.ua = UserAgent()
        
        # Proxy list with authentication
        self.proxies_auth = [
            "http://ce42509537a8eed9520a:25ea7a1cfec42236@gw.dataimpulse.com:10000",
            "http://ce42509537a8eed9520a:25ea7a1cfec42236@gw.dataimpulse.com:10001",
            "http://ce42509537a8eed9520a:25ea7a1cfec42236@gw.dataimpulse.com:10002",
            "http://ce42509537a8eed9520a:25ea7a1cfec42236@gw.dataimpulse.com:10003",
            "http://ce42509537a8eed9520a:25ea7a1cfec42236@gw.dataimpulse.com:10004",
            "http://ce42509537a8eed9520a:25ea7a1cfec42236@gw.dataimpulse.com:10005",
            "http://ce42509537a8eed9520a:25ea7a1cfec42236@gw.dataimpulse.com:10006",
            "http://ce42509537a8eed9520a:25ea7a1cfec42236@gw.dataimpulse.com:10007",
            "http://ce42509537a8eed9520a:25ea7a1cfec42236@gw.dataimpulse.com:10008",
            "http://ce42509537a8eed9520a:25ea7a1cfec42236@gw.dataimpulse.com:10009"
        ]
        
        # Create HTML directory if saving HTML
        if self.save_html:
            os.makedirs("html_pages", exist_ok=True)
        
        print("✅ EnhancedUltraFastBulkDataExtraction initialized with proxy support")
        
    def get_random_proxy_and_headers(self):
        """Get random proxy and headers"""
        proxy = random.choice(self.proxies_auth)
        proxies = {"http": proxy, "https": proxy}
        
        headers = {
            "Connection": "keep-alive",
            "User-Agent": self.ua.random,
            "Accept-Language": "en-US,en;q=0.9",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        }
        
        return proxies, headers, proxy
        
    def get_current_excel_row_count(self):
        """Get current row count from existing Excel file"""
        try:
            if os.path.exists(self.output_file):
                wb = load_workbook(self.output_file, read_only=True)
                ws = wb.active
                row_count = ws.max_row - 1  # Subtract 1 for header row
                wb.close()
                print(f"📊 Found existing Excel file with {row_count} rows")
                return row_count
            return 0
        except Exception as e:
            print(f"❌ Error reading Excel file: {str(e)}")
            return 0

    def get_soup_with_proxy(self, url, page_num=None):
        """Get soup object with proxy rotation and error handling"""
        max_retries = 3
        
        for attempt in range(max_retries):
            try:
                proxies, headers, proxy_url = self.get_random_proxy_and_headers()
                
                response = requests.get(url, headers=headers, proxies=proxies, timeout=20)
                
                if response.status_code == 404:
                    print(f"❌ 404 Not Found: {url}")
                    return None
                
                response.raise_for_status()
                
                # Save HTML if enabled
                if self.save_html and page_num is not None:
                    filename = f"html_pages/page_{page_num}.html"
                    with open(filename, "w", encoding="utf-8") as f:
                        f.write(response.text)
                    print(f"💾 Saved HTML ({len(response.text)} chars) | Proxy: {proxy_url.split('@')[-1]}")
                
                return BeautifulSoup(response.content, 'html.parser')
                
            except requests.exceptions.RequestException as e:
                print(f"⚠️ Attempt {attempt + 1} failed for {url}: {e}")
                if attempt < max_retries - 1:
                    time.sleep(random.uniform(2, 5))
                continue
            except Exception as e:
                print(f"❌ Unexpected error for {url}: {e}")
                break
                
        return None

    def extract_single_url_data(self, url, page_num=None):
        """Extract all data for a single URL with proxy support"""
        soup = self.get_soup_with_proxy(url, page_num)
        if not soup:
            return None
            
        try:
            data = {
                'url': url,
                'prescription': self.prescription(soup),
                'salt_composition': self.salt_compo(soup),
                'side_effects': self.side_effects(soup),
                'product_description': self.product_description(soup),
                'faqs': self.format_faqs_for_excel(self.faqs(soup)),
                'how_drug_works': self.format_how_drug_works_for_excel(self.how_drug_works(soup)),
                'drug_interactions': self.format_drug_interactions_for_excel(self.drug_interaction(soup)),
                'how_to_use': self.format_how_to_use_for_excel(self.how_to_use(soup)),
                'safety_advice': self.format_safety_advice_for_excel(self.safety_advice(soup))
            }
            
            # Add anti-blocking delay
            time.sleep(random.uniform(2, 6))
            
            return data
        except Exception as e:
            print(f"❌ Error extracting data from {url}: {str(e)}")
            return None

    def prescription(self, soup):
        try:
            outer_div = soup.find("div", class_="DrugHeader__prescription-req___34WVy")
            if outer_div:
                span = outer_div.find("span")
                if span:
                    return span.get_text(strip=True)
        except:
            pass
        return ""

    def salt_compo(self, soup):
        try:
            div_tag = soup.find("div", class_="saltInfo DrugHeader__meta-value___vqYM0")
            if div_tag:
                a_tag = div_tag.find("a")
                if a_tag:
                    return a_tag.get_text(strip=True)
        except:
            pass
        return ""

    def side_effects(self, soup):
        try:
            side_effects_div = soup.find("div", id="side_effects")
            if side_effects_div:
                return side_effects_div.get_text(separator="\n", strip=True)
        except:
            pass
        return ""

    def product_description(self, soup):
        try:
            product_intro_div = soup.find("div", class_="DrugOverview__container___CqA8x")
            if product_intro_div:
                return product_intro_div.get_text(separator=" ", strip=True)
        except:
            pass
        return ""

    def faqs(self, soup):
        faq_list = []
        try:
            faq_div = soup.find("div", id="faq")
            if faq_div:
                for tile in faq_div.find_all("div", class_="Faqs__tile___1B58W"):
                    question_tag = tile.find("h3", class_="Faqs__ques___1iPB9")
                    answer_tag = tile.find("div", class_="Faqs__ans___1uuIW")
                    if question_tag and answer_tag:
                        faq_list.append({
                            "Q": question_tag.get_text(strip=True),
                            "A": answer_tag.get_text(strip=True)
                        })
        except:
            pass
        return faq_list

    def how_drug_works(self, soup):
        try:
            container = soup.find("div", id="how_drug_works")
            if container:
                title_tag = container.find("h2")
                content_tag = container.find("div", class_="DrugOverview__content___22ZBX")
                title = title_tag.get_text(strip=True) if title_tag else ""
                content = content_tag.get_text(strip=True) if content_tag else ""
                return {"title": title, "content": content}
        except:
            pass
        return {}

    def drug_interaction(self, soup):
        try:
            interaction_div = soup.find("div", id="drug_interaction")
            if interaction_div:
                title_tag = interaction_div.find("h2")
                description_tag = interaction_div.find("div", class_="DrugInteraction__desc___2y8bR")
                title = title_tag.get_text(strip=True) if title_tag else ""
                description = description_tag.get_text(strip=True) if description_tag else ""

                interactions = []
                for drug_block in interaction_div.find_all("div", class_="DrugInteraction__drug___1XyzI"):
                    name = drug_block.get_text(strip=True)
                    desc_block = drug_block.find_next("div", class_="DrugInteraction__interaction-text___1hOwx")
                    desc_text = desc_block.get_text(" ", strip=True) if desc_block else ""
                    interactions.append({"drug": name, "description": desc_text})

                return {
                    "title": title,
                    "description": description,
                    "interactions": interactions
                }
        except:
            pass
        return {}

    def how_to_use(self, soup):
        try:
            how_to_use_div = soup.find("div", id="how_to_use")
            if how_to_use_div:
                heading_tag = how_to_use_div.find("h2")
                instructions_tag = how_to_use_div.find("div", class_="DrugOverview__content___22ZBX")
                heading = heading_tag.get_text(strip=True) if heading_tag else ""
                instructions = instructions_tag.get_text(" ", strip=True) if instructions_tag else ""
                return {"heading": heading, "instructions": instructions}
        except:
            pass
        return {}

    def safety_advice(self, soup):
        try:
            safety_div = soup.find("div", id="safety_advice")
            if not safety_div:
                return {}

            heading_tag = safety_div.find("h2")
            heading = heading_tag.get_text(strip=True) if heading_tag else ""

            safety_items = []
            warnings = safety_div.find_all("div", class_="DrugOverview__warning-top___UD3xX")

            for warning in warnings:
                label_tag = warning.find("span")
                status_tag = warning.find("div", class_="DrugOverview__warning-tag___aHZlc")

                label = label_tag.get_text(strip=True) if label_tag else ""
                status = status_tag.get_text(strip=True) if status_tag else ""

                desc_div = warning.find_next_sibling("div", class_="DrugOverview__content___22ZBX")
                description = desc_div.get_text(" ", strip=True) if desc_div else ""

                safety_items.append({
                    "label": label,
                    "status": status,
                    "description": description
                })

            return {"heading": heading, "items": safety_items}
        except:
            pass
        return {}

    def format_faqs_for_excel(self, faqs):
        if not faqs:
            return ""
        formatted_text = ""
        for i, faq in enumerate(faqs, 1):
            formatted_text += f"Q{i}: {faq['Q']}\n"
            formatted_text += f"A{i}: {faq['A']}\n\n"
        return formatted_text.strip()

    def format_drug_interactions_for_excel(self, interactions_data):
        if not interactions_data:
            return ""
        formatted_text = ""
        if interactions_data.get('title'):
            formatted_text += f"TITLE: {interactions_data['title']}\n\n"
        if interactions_data.get('description'):
            formatted_text += f"DESCRIPTION: {interactions_data['description']}\n\n"
        if interactions_data.get('interactions'):
            formatted_text += "DRUG INTERACTIONS:\n"
            for i, interaction in enumerate(interactions_data['interactions'], 1):
                formatted_text += f"{i}. {interaction['drug']}\n"
                formatted_text += f"   Description: {interaction['description']}\n\n"
        return formatted_text.strip()

    def format_safety_advice_for_excel(self, safety_data):
        if not safety_data:
            return ""
        formatted_text = ""
        if safety_data.get('heading'):
            formatted_text += f"HEADING: {safety_data['heading']}\n\n"
        if safety_data.get('items'):
            for i, item in enumerate(safety_data['items'], 1):
                formatted_text += f"{i}. {item['label']}\n"
                formatted_text += f"   Status: {item['status']}\n"
                formatted_text += f"   Description: {item['description']}\n\n"
        return formatted_text.strip()

    def format_how_drug_works_for_excel(self, how_works_data):
        if not how_works_data:
            return ""
        formatted_text = ""
        if how_works_data.get('title'):
            formatted_text += f"TITLE: {how_works_data['title']}\n\n"
        if how_works_data.get('content'):
            formatted_text += f"CONTENT: {how_works_data['content']}"
        return formatted_text.strip()

    def format_how_to_use_for_excel(self, how_to_use_data):
        if not how_to_use_data:
            return ""
        formatted_text = ""
        if how_to_use_data.get('heading'):
            formatted_text += f"HEADING: {how_to_use_data['heading']}\n\n"
        if how_to_use_data.get('instructions'):
            formatted_text += f"INSTRUCTIONS: {how_to_use_data['instructions']}"
        return formatted_text.strip()

    def create_excel_with_headers(self):
        """Create Excel file with headers if it doesn't exist"""
        if os.path.exists(self.output_file):
            print(f"📊 Excel file already exists: {self.output_file}")
            return
            
        wb = Workbook()
        ws = wb.active
        ws.title = "Drug Data"

        headers = [
            "URL",
            "Prescription Required",
            "Salt Composition", 
            "Side Effects",
            "Product Description",
            "FAQs",
            "How Drug Works",
            "Drug Interactions",
            "How to Use",
            "Safety Advice"
        ]
        
        ws.append(headers)
        
        # Format headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        column_widths = [50, 20, 30, 40, 60, 60, 60, 60, 60, 60]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        wb.save(self.output_file)
        print(f"✅ Created Excel file: {self.output_file}")

    def create_csv_with_headers(self, csv_file):
        """Create CSV file with headers if it doesn't exist"""
        if os.path.exists(csv_file):
            print(f"📊 CSV file already exists: {csv_file}")
            return
        headers = [
            "URL",
            "Prescription Required",
            "Salt Composition", 
            "Side Effects",
            "Product Description",
            "FAQs",
            "How Drug Works",
            "Drug Interactions",
            "How to Use",
            "Safety Advice"
        ]
        with open(csv_file, mode='w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
        print(f"✅ Created CSV file: {csv_file}")

    def append_data_to_csv_batch(self, data_batch, csv_file):
        """Append batch of data to CSV file"""
        with self.lock:
            try:
                with open(csv_file, mode='a', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    for data in data_batch:
                        if data:
                            row_data = [
                                data['url'],
                                data['prescription'],
                                data['salt_composition'],
                                data['side_effects'],
                                data['product_description'],
                                data['faqs'],
                                data['how_drug_works'],
                                data['drug_interactions'],
                                data['how_to_use'],
                                data['safety_advice']
                            ]
                            writer.writerow(row_data)
                print(f"✅ Saved batch to CSV. Total rows written: {self.processed_count}")
            except Exception as e:
                print(f"❌ Error saving to CSV: {str(e)}")

    def process_urls_sequential(self, urls, start_from=0):
        """Process URLs sequentially with proxy rotation (like the smaller script)"""
        results = []
        
        for i, url in enumerate(urls):
            actual_index = start_from + i
            try:
                print(f"\n🔄 Processing [{actual_index}]: {url}")
                
                data = self.extract_single_url_data(url.strip(), actual_index)
                if data:
                    results.append(data)
                    with self.lock:
                        self.processed_count += 1
                        print(f"✅ [{actual_index}] Successfully processed")
                else:
                    with self.lock:
                        self.failed_urls.append(url)
                        print(f"❌ [{actual_index}] Failed to process")
                
                # Anti-blocking sleep with longer delays
                if actual_index % 50 == 0 and actual_index > 0:
                    sleep_time = random.uniform(8, 15)
                    print(f"😴 Long pause ({sleep_time:.1f}s) after 50 requests...")
                    time.sleep(sleep_time)
                else:
                    sleep_time = random.uniform(2, 6)
                    time.sleep(sleep_time)
                    
            except Exception as e:
                with self.lock:
                    self.failed_urls.append(url)
                    print(f"⚠️ Error processing [{actual_index}] {url}: {str(e)}")
                time.sleep(random.uniform(3, 8))
                
        return results

    def enhanced_process(self, start_from=0, limit=None, csv_file=None, use_threading=False, max_workers=3):
        """Enhanced processing with proxy support"""
        print(f"🚀 Starting enhanced processing from row {start_from + 1}")

        # Read CSV efficiently
        try:
            print(f"📊 Reading CSV file: {self.csv_file_path}")
            df = pd.read_csv(self.csv_file_path)
            
            # Use specified column index for URLs
            if self.column_index >= len(df.columns):
                print(f"❌ Column index {self.column_index} out of range. Available columns: {len(df.columns)}")
                return
                
            all_urls = df.iloc[:, self.column_index].dropna().tolist()
            
            # Apply start_from and limit
            if limit:
                urls = all_urls[start_from:start_from+limit]
            else:
                urls = all_urls[start_from:]
                
            total_urls = len(urls)
            print(f"📈 Loaded {total_urls} URLs to process (starting from row {start_from + 1})")
        except Exception as e:
            print(f"❌ Error reading CSV: {str(e)}")
            return

        # Set default CSV file name
        if not csv_file:
            csv_file = f"enhanced_drug_data_{start_from}_{total_urls}.csv"

        # Create CSV file
        self.create_csv_with_headers(csv_file)

        print(f"⚡ Processing with proxy rotation and anti-blocking delays")
        start_time = time.time()

        # Process URLs
        if use_threading:
            print(f"🧵 Using threading with {max_workers} workers")
            # For threading, we'll use smaller batches
            batch_size = 20
            pending_data = []
            
            for i in range(0, total_urls, batch_size):
                batch_urls = urls[i:i+batch_size]
                batch_results = []
                
                # Process batch sequentially to maintain proxy rotation and delays
                for j, url in enumerate(batch_urls):
                    data = self.extract_single_url_data(url.strip(), start_from + i + j)
                    if data:
                        batch_results.append(data)
                        self.processed_count += 1
                    else:
                        self.failed_urls.append(url)
                
                pending_data.extend(batch_results)
                
                # Save every 50 records
                if len(pending_data) >= 50:
                    self.append_data_to_csv_batch(pending_data, csv_file)
                    pending_data = []
        else:
            print("🔄 Using sequential processing")
            results = self.process_urls_sequential(urls, start_from)
            
            # Save all results
            if results:
                self.append_data_to_csv_batch(results, csv_file)

        # Final statistics
        total_time = time.time() - start_time
        print("\n" + "="*60)
        print("🎉 PROCESSING COMPLETE!")
        print(f"📊 Total URLs processed: {self.processed_count}")
        print(f"❌ Failed URLs: {len(self.failed_urls)}")
        if self.processed_count + len(self.failed_urls) > 0:
            success_rate = (self.processed_count/(self.processed_count + len(self.failed_urls)))*100
            print(f"✅ Success rate: {success_rate:.1f}%")
        print(f"⏱️ Total time: {total_time/60:.1f} minutes")
        if total_time > 0:
            print(f"⚡ Average speed: {self.processed_count/total_time:.2f} URLs/second")
        print(f"💾 CSV file: {csv_file}")

        # Save failed URLs
        if self.failed_urls:
            failed_df = pd.DataFrame(self.failed_urls, columns=['failed_urls'])
            failed_file = f"failed_urls_from_{start_from}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            failed_df.to_csv(failed_file, index=False)
            print(f"📝 Failed URLs saved to: {failed_file}")

    def csv_to_excel(self, csv_file, excel_file):
        """Convert CSV file to Excel file"""
        df = pd.read_csv(csv_file)
        df.to_excel(excel_file, index=False)
        print(f"✅ Converted {csv_file} to {excel_file}")

# Main execution
if __name__ == "__main__":
    print("🚀 Starting Enhanced UltraFast Bulk Data Extraction with Proxy Support...")
    
    # Configuration
    csv_file = "pain_relief.csv"  # Update this to your CSV file path
    column_index = 5  # Column index for URLs (0-based, so 5 means 6th column)
    output_csv = "enhanced_drug_data.csv"
    output_excel = "enhanced_drug_data.xlsx"
    
    # Check if CSV file exists
    if not os.path.exists(csv_file):
        print(f"❌ CSV file not found: {csv_file}")
        print("Please make sure the CSV file is in the same directory as this script.")
        exit(1)

    # Initialize extractor
    extractor = EnhancedUltraFastBulkDataExtraction(
        csv_file_path=csv_file,
        column_index=column_index,
        save_html=True  # Set to False if you don't want to save HTML files
    )

    # Start processing
    extractor.enhanced_process(
        start_from=0,      # Start from first row
        limit=10,          # Process only 10 URLs for testing (remove or set to None for all)
        csv_file=output_csv,
        use_threading=False,  # Set to True for threading (not recommended with proxy rotation)
        max_workers=3
    )

    # Convert to Excel after processing
    extractor.csv_to_excel(output_csv, output_excel)